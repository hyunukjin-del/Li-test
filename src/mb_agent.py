import os
import sys
import json
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 화학 분자량 상수 (g/mol)
MW_LI2CO3 = 73.89
MW_CAO = 56.08
MW_CAOH2 = 74.09
MW_LIOH = 23.95
MW_CACO3 = 100.09
MW_LI = 6.941

def run_pipeline(json_path: str):
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    run_no = data["run_no"]
    raw = data["raw_materials"]
    filt = data["filtration"]
    wash = data["washing"]
    calc = data["calcination"]
    icp = data.get("icp_analysis", {})

    # 1. 화학양론 및 몰수 계산
    n_li2co3 = raw["li2co3_mass"] / MW_LI2CO3
    total_cao_in = raw["fresh_cao_mass"] + raw["recycled_cao_mass"]
    n_cao = (total_cao_in * raw["cao_purity"]) / MW_CAO
    limiting = "Li2CO3" if n_li2co3 <= n_cao else "CaO"
    excess_pct = (n_cao / n_li2co3 - 1.0) * 100
    theo_lioh_mass = (n_li2co3 * 2.0) * MW_LIOH
    theo_caco3_mass = n_li2co3 * MW_CACO3

    # 2. 질량 수지 (M/B)
    total_in = raw["li2co3_mass"] + raw["li2co3_water"] + total_cao_in + raw["slurry_water"]
    total_out = filt["primary_filtrate_mass"] + filt["wet_cake_mass"]
    loss_mass = total_in - total_out
    mass_closure = (total_out / total_in) * 100.0
    cake_moisture = (1.0 - (filt["sample_dry_mass"] / filt["sample_wet_mass"])) * 100.0
    est_total_dry_solids = filt["wet_cake_mass"] * (filt["sample_dry_mass"] / filt["sample_wet_mass"])

    # 3. 소성 및 순도 분석
    loi_pct = ((calc["test_dry_cake_mass"] - calc["calcined_cao_mass"]) / calc["test_dry_cake_mass"]) * 100.0
    cao_yield_dry = (calc["calcined_cao_mass"] / calc["test_dry_cake_mass"]) * 100.0
    purity_caco3 = max(0.0, min(100.0, ((loi_pct/100.0) - 0.2432) / (0.4397 - 0.2432) * 100.0))
    pot_total_cao = est_total_dry_solids * (cao_yield_dry / 100.0)
    ca_loop_recovery = (pot_total_cao / total_cao_in) * 100.0

    # 4. ICP 기반 Li 회수율
    li_in_g = n_li2co3 * 2 * MW_LI
    v_primary_l = (filt["primary_filtrate_mass"] / filt["primary_filtrate_sg"]) / 1000.0
    v_wash_l = (wash["wash_sol_mass"] / wash["wash_sol_sg"]) / 1000.0
    li_p_g = (icp.get("primary_li_mg_l", 0.0) * v_primary_l) / 1000.0
    li_w_g = (icp.get("wash_li_mg_l", 0.0) * v_wash_l) / 1000.0
    total_li_rec_g = li_p_g + li_w_g
    li_recovery_pct = (total_li_rec_g / li_in_g) * 100.0 if icp.get("primary_li_mg_l", 0) > 0 else 0.0

    # 5. 차기 회차 가이드
    target_cao = 92.42
    recycled_sample_cao = calc["calcined_cao_mass"]
    fresh_makeup = max(0.0, target_cao - recycled_sample_cao)

    # 6. AI 진단 도출
    diagnostics = []
    if loss_mass > 50.0:
        diagnostics.append(f"반응 중 수분 {loss_mass:.1f}g 증발 손실 발생 (농도 유지를 위해 환류장치 장착 또는 보충수 설정 권장)")
    if wash["wash_sol_ph"] >= 13.5:
        diagnostics.append(f"수세액 pH({wash['wash_sol_ph']}) 강알칼리 확인 -> 케이크 내부 잔류 LiOH 농축액이 세척수로 성공 회수됨")
    if loi_pct < 43.0:
        diagnostics.append(f"하소 감율({loi_pct:.1f}%) 편차 -> 잉여 Ca(OH)2 약 {100-purity_caco3:.1f}% 공침/잔류 확인")
    if calc["temp_c"] >= 1000.0:
        diagnostics.append(f"1000℃ 고온 소성에 따른 소결(Sintering) 주의 -> Run {run_no+1} 슬러리 소화 교반시간 +15분 권장")

    # 7. 엑셀 파일 생성 (4개 시트)
    wb = Workbook()
    ws_raw = wb.active
    ws_raw.title = "1_원본입력"
    ws_calc = wb.create_sheet(title="2_MB계산결과")
    ws_trend = wb.create_sheet(title="3_트렌드누적")
    ws_issue = wb.create_sheet(title="4_이슈및진단")

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="맑은 고딕", size=10)

    # [Sheet 1: 원본 입력]
    ws_raw.append(["구분", "항목", "입력값", "단위"])
    ws_raw.append(["투입원료", "Li2CO3 투입량", raw["li2co3_mass"], "g"])
    ws_raw.append(["투입원료", "Li2CO3 용매수", raw["li2co3_water"], "g"])
    ws_raw.append(["투입원료", "신품 CaO 투입량", raw["fresh_cao_mass"], "g"])
    ws_raw.append(["투입원료", "재생 CaO 투입량", raw["recycled_cao_mass"], "g"])
    ws_raw.append(["투입원료", "슬러리 조제수", raw["slurry_water"], "g"])
    ws_raw.append(["1차여과", "1차 여액 무게", filt["primary_filtrate_mass"], "g"])
    ws_raw.append(["1차여과", "여액 비중 / pH", f"{filt['primary_filtrate_sg']} / {filt['primary_filtrate_ph']}", "-"])
    ws_raw.append(["1차여과", "1차 습케이크", filt["wet_cake_mass"], "g"])
    ws_raw.append(["케익수세", "수세 후 건조케이크", wash["washed_dry_cake_mass"], "g"])
    ws_raw.append(["케익수세", "수세 여액 무게", wash["wash_sol_mass"], "g"])
    ws_raw.append(["소성재생", "소성 투입 건조케익", calc["test_dry_cake_mass"], "g"])
    ws_raw.append(["소성재생", "소성 회수 CaO", calc["calcined_cao_mass"], "g"])

    # [Sheet 2: M/B 계산 결과]
    ws_calc.append(["대분류", "지표명", "계산값", "단위", "평가/비고"])
    ws_calc.append(["화학양론", "제한반응물", limiting, "-", f"Li2CO3 {n_li2co3:.2f}mol vs CaO {n_cao:.2f}mol"])
    ws_calc.append(["화학양론", "이론 LiOH 생성량", round(theo_lioh_mass, 2), "g", "100% 전환 기준"])
    ws_calc.append(["물질수지", "M/B 정합성(Closure)", round(mass_closure, 2), "%", f"손실(증발): {loss_mass:.1f}g"])
    ws_calc.append(["물질수지", "1차 케이크 함수율", round(cake_moisture, 2), "%", "건조 전 기준"])
    ws_calc.append(["소성/재생", "실측 하소 감율(LOI)", round(loi_pct, 2), "%", f"CaCO3 순도 약 {purity_caco3:.1f}%"])
    ws_calc.append(["소성/재생", "Ca 루프 회수율", round(ca_loop_recovery, 2), "%", "원소 기준"])
    ws_calc.append(["ICP/수율", "총 Li 회수율", round(li_recovery_pct, 2), "%", f"1차여액 {li_p_g:.2f}g + 수세액 {li_w_g:.2f}g"])
    ws_calc.append(["차기투입", "Run n+1 신품 CaO 보충량", round(fresh_makeup, 2), "g", f"재생분 {recycled_sample_cao:.2f}g 활용"])

    # [Sheet 4: 이슈 및 진단]
    ws_issue.append(["No", "AI 공정 진단 및 조치 가이드"])
    for idx, diag in enumerate(diagnostics, 1):
        ws_issue.append([idx, diag])

    for ws in [ws_raw, ws_calc, ws_issue]:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = data_font
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    os.makedirs("reports", exist_ok=True)
    report_file = f"reports/MB_Report_Run_{run_no:03d}.xlsx"
    wb.save(report_file)

    # 8. CSV 트렌드 파일 누적 업데이트
    os.makedirs("data", exist_ok=True)
    trend_csv = "data/trend_history.csv"
    row_data = {
        "Run_No": run_no,
        "Li2CO3_in_g": raw["li2co3_mass"],
        "Fresh_CaO_g": raw["fresh_cao_mass"],
        "Recycled_CaO_g": raw["recycled_cao_mass"],
        "MB_Closure_pct": round(mass_closure, 2),
        "Li_Recovery_pct": round(li_recovery_pct, 2),
        "Calcination_LOI_pct": round(loi_pct, 2),
        "Ca_Loop_Yield_pct": round(ca_loop_recovery, 2),
        "Next_Fresh_Makeup_g": round(fresh_makeup, 2)
    }
    df_new = pd.DataFrame([row_data])
    if os.path.exists(trend_csv):
        df_old = pd.read_csv(trend_csv)
        df_all = pd.concat([df_old[df_old["Run_No"] != run_no], df_new]).sort_values("Run_No")
    else:
        df_all = df_new
    df_all.to_csv(trend_csv, index=False)

    # 9. GitHub Step Summary용 Markdown 작성
    summary_md = f"""### 📊 습식반응 M/B 자동화 Agent 리포트 (Run {run_no})

| 핵심 KPI | 계산값 | 비고 |
| :--- | :---: | :--- |
| **제한반응물 / 투입 몰비** | `{limiting}` | Li₂CO₃ {n_li2co3:.2f}mol vs CaO {n_cao:.2f}mol ({excess_pct:+.1f}%) |
| **Mass Balance Closure** | **{mass_closure:.2f}%** | 증발 손실 {loss_mass:.1f}g |
| **하소 감율 (LOI)** | **{loi_pct:.2f}%** | 건조 케이크 내 CaCO₃ 순도 ~{purity_caco3:.1f}% |
| **총 Li 회수율 (ICP)** | **{li_recovery_pct:.2f}%** | 1차 여액 + 수세액 합산 |
| **Ca Loop 원소 회수율** | **{ca_loop_recovery:.2f}%** | 재생 잠재 CaO {pot_total_cao:.1f}g |

#### 🤖 AI 공정 진단 및 체크포인트
"""
    for d in diagnostics:
        summary_md += f"- {d}\n"

    summary_md += f"""
#### 📋 차기 회차 (Run {run_no+1}) 추천 배합비
- **재생 CaO 사용량:** `{recycled_sample_cao:.2f} g`
- **신품 CaO 보충량 (Make-up):** `{fresh_makeup:.2f} g`
- **슬러리 물 투입량:** `{raw['slurry_water']} g`

> 📁 **생성된 엑셀 리포트:** `{report_file}`
"""
    with open("summary.md", "w", encoding="utf-8") as f:
        f.write(summary_md)

    print(f"Run {run_no} M/B 연산 및 리포트 생성이 완료되었습니다.")

if __name__ == "__main__":
    target_json = sys.argv[1] if len(sys.argv) > 1 else "data/runs/run_001.json"
    run_pipeline(target_json)
