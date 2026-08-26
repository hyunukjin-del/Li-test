import streamlit as st
import pandas as pd
import numpy as np
import io
import json
import smtplib
from datetime import datetime
from PIL import Image
import google.generativeai as genai
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# 화학 분자량 상수 (g/mol)
MW_LI2CO3 = 73.89
MW_CAO = 56.08
MW_CAOH2 = 74.09
MW_LIOH = 23.95
MW_CACO3 = 100.09
MW_LI = 6.941
MW_CA = 40.08
MW_NA = 22.99
MW_SI = 28.09
MW_MG = 24.31
MW_K = 39.10

AGENT_TITLE = "LC-LH전환반응 M/B자동화 및 거동예측 Agent tool"

st.set_page_config(page_title=AGENT_TITLE, page_icon="🧪", layout="wide")

# --------------------------------------------------------------------------
# [1] 기본 세션 상태 및 변수 초기화
# --------------------------------------------------------------------------
DEFAULT_DATA = {
    "run_no": 1,
    "li2co3_mass": 95.34,
    "li2co3_water": 1040.0,
    "fresh_cao_mass": 92.42,
    "recycled_cao_mass": 0.0,
    "slurry_water": 831.0,
    "temp_c": 80.0,
    "time_h": 2.0,
    "primary_filtrate_mass": 1646.0,
    "primary_filtrate_sg": 1.035,
    "primary_filtrate_ph": 12.81,
    "wet_cake_mass": 311.0,
    "sample_wet": 27.7,
    "sample_dry": 14.8,
    "wash_sol_mass": 832.0,
    "wash_sol_sg": 1.000,
    "wash_sol_ph": 13.70,
    "test_dry_cake": 40.6,
    "calcined_cao": 23.9,
    "calc_temp": 1000.0,
    "calc_time": 1.0,
    "icp_li_1": 10500.0,
    "icp_ca_1": 120.0,
    "icp_na_1": 45.0,
    "icp_si_1": 8.5,
    "icp_mg_1": 1.2,
    "icp_k_1": 15.0,
    "icp_li_w": 1400.0,
    "icp_ca_w": 80.0,
    "icp_na_w": 6.0,
    "icp_si_w": 2.1,
    "icp_mg_w": 0.3,
    "icp_k_w": 2.0,
}

for k, v in DEFAULT_DATA.items():
    if k not in st.session_state:
        st.session_state[k] = v

if "gemini_api_key" not in st.session_state:
    st.session_state.gemini_api_key = ""
if "email_recipients" not in st.session_state:
    st.session_state.email_recipients = "user@company.com"
if "email_sender" not in st.session_state:
    st.session_state.email_sender = "sender@gmail.com"
if "email_password" not in st.session_state:
    st.session_state.email_password = ""
if "smtp_server" not in st.session_state:
    st.session_state.smtp_server = "smtp.gmail.com"
if "smtp_port" not in st.session_state:
    st.session_state.smtp_port = 587
if "auto_email_on_save" not in st.session_state:
    st.session_state.auto_email_on_save = True
if "email_logs" not in st.session_state:
    st.session_state.email_logs = []

if "history" not in st.session_state:
    st.session_state.history = pd.DataFrame([
        {
            "회차 (Run)": 1, 
            "구분": "실측치 (Actual)",
            "Li 회수율 (%)": 95.80, 
            "1차여액 Li농도 (mg/L)": 10500.0,
            "1차여액 LiOH농도 (g/L)": round(10500.0 * (MW_LIOH / MW_LI) / 1000, 2),
            "M/B 닫힘율 (%)": 95.06, 
            "하소 감율 LOI (%)": 41.13, 
            "CaO 활성도 (%)": 100.0,
            "신품 CaO 보충량 (g)": 68.52
        }
    ])

if "chat_messages" not in st.session_state:
    st.session_state.chat_messages = [
        {"role": "assistant", "content": f"안녕하세요! **{AGENT_TITLE}**입니다. Google Gemini Vision 기반 사진 인식, M/B 연산, 거동예측에 대해 무엇이든 질문해 주세요."}
    ]

# --------------------------------------------------------------------------
# [2] Google Gemini Vision OCR 분석 함수 (무료 티어 적용)
# --------------------------------------------------------------------------
def parse_image_with_vision(image_bytes, doc_type="lab_note"):
    api_key = st.session_state.gemini_api_key.strip()
    if not api_key:
        return None, "사이드바에 Google Gemini API Key를 먼저 입력해 주세요. (무료 발급 가능)"

    try:
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel("gemini-1.5-flash")
        img = Image.open(io.BytesIO(image_bytes))

        if doc_type == "lab_note":
            prompt = """이 이미지는 탄산리튬(LC) 가성화 및 Ca-Loop 습식 공정의 수기 실험 노트 또는 기록지입니다.
이미지에 적힌 수치를 정확히 읽어내어 아래 JSON 포맷으로만 반환해 주세요.
단위(g, mL, ℃ 등)는 제외하고 순수 숫자(float)만 넣어주세요. 판독할 수 없는 항목은 null로 설정하세요.

{
  "run_no": number,
  "li2co3_mass": number,
  "li2co3_water": number,
  "fresh_cao_mass": number,
  "recycled_cao_mass": number,
  "slurry_water": number,
  "temp_c": number,
  "time_h": number,
  "primary_filtrate_mass": number,
  "primary_filtrate_sg": number,
  "primary_filtrate_ph": number,
  "wet_cake_mass": number,
  "sample_wet": number,
  "sample_dry": number,
  "wash_sol_mass": number,
  "wash_sol_sg": number,
  "wash_sol_ph": number,
  "test_dry_cake": number,
  "calcined_cao": number
}"""
        else: # icp_report
            prompt = """이 이미지는 용액 ICP 분석 기기 화면 또는 시험 성적서 인쇄물입니다.
1차 여액(Primary Filtrate)과 수세액(Wash Solution)의 Li, Ca, Na, Si, Mg, K 농도(mg/L)를 추출하여 아래 JSON 포맷으로만 반환해 주세요.
단위는 제외하고 순수 숫자만 넣어주세요. 판독할 수 없는 항목은 null로 설정하세요.

{
  "icp_li_1": number, "icp_ca_1": number, "icp_na_1": number,
  "icp_si_1": number, "icp_mg_1": number, "icp_k_1": number,
  "icp_li_w": number, "icp_ca_w": number, "icp_na_w": number,
  "icp_si_w": number, "icp_mg_w": number, "icp_k_w": number
}"""

        response = model.generate_content(
            [img, prompt],
            generation_config={"response_mime_type": "application/json"}
        )
        result_json = json.loads(response.text)
        return result_json, None
    except Exception as e:
        return None, str(e)

# --------------------------------------------------------------------------
# [3] 이메일 리포트 발송 공통 함수
# --------------------------------------------------------------------------
def send_email_report(run_num, mass_cls, loss_m, li_rec_tot, li_rec_1, li_rec_w, 
                      lioh_conc, li_1, ca_1, na_1, si_1, mg_1, k_1, 
                      loi, purity, makeup, cao_rec, df_icp_tbl, df_sim_tbl, is_auto=True):
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    recipients = [r.strip() for r in st.session_state.email_recipients.split(",") if r.strip()]
    
    if not recipients:
        return False, "수신자 이메일 주소가 비어 있습니다."

    sender = st.session_state.email_sender.strip()
    pw = st.session_state.email_password.strip().replace(" ", "")
    smtp_host = st.session_state.smtp_server.strip()
    port_num = int(st.session_state.smtp_port)

    if not sender or not pw:
        return False, "발신자 계정 또는 비밀번호가 비어 있습니다. (5번 탭 확인)"

    try:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "MB_종합결과"
        ws1.append(["지표명", "수치", "단위", "평가"])
        ws1.append(["실험 회차", run_num, "Run", "정상"])
        ws1.append(["총 Li 회수율(ICP)", round(li_rec_tot, 2), "%", "1차여액+수세액"])
        ws1.append(["1차 여액 LiOH 농도", round(lioh_conc, 2), "g/L", f"Li: {li_1:.1f} mg/L"])
        ws1.append(["M/B 정합성", round(mass_cls, 2), "%", f"증발 {loss_m:.1f}g"])
        ws1.append(["소성 감율(LOI)", round(loi, 2), "%", f"CaCO3 순도 ~{purity:.1f}%"])
        ws1.append(["차기 신품 CaO 보충량", round(makeup, 2), "g", f"재생분 {cao_rec:.1f}g"])

        if df_icp_tbl is not None and not df_icp_tbl.empty:
            ws2 = wb.create_sheet(title="ICP_원소분석")
            ws2.append(list(df_icp_tbl.columns))
            for _, r in df_icp_tbl.iterrows():
                ws2.append(list(r.values))

        if df_sim_tbl is not None and not df_sim_tbl.empty:
            ws3 = wb.create_sheet(title="트렌드시뮬레이션")
            ws3.append(list(df_sim_tbl.columns))
            for _, r in df_sim_tbl.iterrows():
                ws3.append(list(r.values))

        excel_buf = io.BytesIO()
        wb.save(excel_buf)
        excel_buf.seek(0)
        file_name = f"MB_Report_Run_{run_num:03d}.xlsx"

        mail_subject = f"[{AGENT_TITLE}] Run {run_num} M/B 및 ICP 분석 종합 리포트"
        msg = MIMEMultipart()
        msg["From"] = sender
        msg["To"] = ", ".join(recipients)
        msg["Subject"] = mail_subject

        html_body = f"""
        <h3>🧪 {AGENT_TITLE} - Run {run_num} 자동 리포트</h3>
        <hr>
        <h4>📊 핵심 KPI 요약</h4>
        <ul>
            <li><b>총 Li 회수율:</b> <span style="color:#0284C7; font-weight:bold;">{li_rec_tot:.2f}%</span> (1차 여액: {li_rec_1:.2f}%, 수세액: {li_rec_w:.2f}%)</li>
            <li><b>1차 여액 LiOH 농도:</b> {lioh_conc:.2f} g/L (Li: {li_1:,.1f} mg/L)</li>
            <li><b>M/B 정합성(Closure):</b> {mass_cls:.2f}% (증발 손실: {loss_m:.1f}g)</li>
            <li><b>소성 감율(LOI):</b> {loi:.2f}% (CaCO₃ 추정 순도: {purity:.1f}%)</li>
            <li><b>차기(Run {run_num+1}) 신품 CaO 보충량:</b> {makeup:.2f}g</li>
        </ul>
        <h4>🧪 주요 불순물 농도 (1차 여액)</h4>
        <ul>
            <li>Ca: {ca_1} mg/L | Na: {na_1} mg/L | Si: {si_1} mg/L | Mg: {mg_1} mg/L | K: {k_1} mg/L</li>
        </ul>
        <p>※ 세부 분석 데이터 엑셀 파일(<b>{file_name}</b>)을 첨부하였습니다.</p>
        """
        msg.attach(MIMEText(html_body, "html", "utf-8"))

        part = MIMEApplication(excel_buf.read(), Name=file_name)
        part['Content-Disposition'] = f'attachment; filename="{file_name}"'
        msg.attach(part)

        if port_num == 465:
            server = smtplib.SMTP_SSL(smtp_host, port_num, timeout=15)
        else:
            server = smtplib.SMTP(smtp_host, port_num, timeout=15)
            server.starttls()
            
        server.login(sender, pw)
        server.send_message(msg)
        server.quit()

        st.session_state.email_logs.append({
            "발송일시": now_str, "회차 (Run)": f"Run {run_num}", "수신자": ", ".join(recipients),
            "메일 제목": mail_subject, "발송 상태": "✅ 성공", "첨부 파일": file_name,
            "비고": "자동 발송" if is_auto else "수동 발송"
        })
        return True, f"[{', '.join(recipients)}]로 리포트 메일이 발송되었습니다!"
    except Exception as e:
        err_msg = str(e)
        st.session_state.email_logs.append({
            "발송일시": now_str, "회차 (Run)": f"Run {run_num}", "수신자": ", ".join(recipients),
            "메일 제목": f"[{AGENT_TITLE}] Run {run_num} 리포트", "발송 상태": "❌ 실패",
            "첨부 파일": "-", "비고": f"SMTP 오류: {err_msg}"
        })
        return False, f"메일 발송 실패: {err_msg}"

# --------------------------------------------------------------------------
# [4] 사이드바: Google Gemini API Key 설정 (무료)
# --------------------------------------------------------------------------
with st.sidebar:
    st.header("🔑 Google Gemini AI 설정")
    st.caption("Google AI Studio의 무료 API Key를 입력하세요. (신용카드 불필요)")
    st.session_state.gemini_api_key = st.text_input(
        "Google Gemini API Key", 
        value=st.session_state.gemini_api_key, 
        type="password",
        help="aistudio.google.com에서 발급받은 AIzaSy... 키를 입력하세요."
    )
    if st.session_state.gemini_api_key:
        st.success("✅ Gemini AI 준비 완료 (무료 티어)")
    else:
        st.info("💡 aistudio.google.com 에서 무료로 발급받은 키를 넣으시면 사진 인식이 활성화됩니다.")
    st.divider()

# --------------------------------------------------------------------------
# [5] 메인 화면 및 탭 구성
# --------------------------------------------------------------------------
st.title(f"🧪 {AGENT_TITLE}")
st.caption("Google Gemini 사진 인식 | M/B 자동 연산 | ICP 성분별 회수율 | 회차별 거동예측 | 리포트 자동 발송")

main_tab1, main_tab2, main_tab3, main_tab4, main_tab5 = st.tabs([
    "1️⃣ 실험 데이터 입력 & M/B 연산", 
    "2️⃣ 🧪 용액 ICP 분석 & 회수율", 
    "3️⃣ 📈 회차별 트렌드 & 거동예측", 
    "4️⃣ 💬 AI 공정 대화창", 
    "5️⃣ 📧 리포트 메일 발송 및 현황"
])

# --------------------------------------------------------------------------
# TAB 1: 실험 데이터 입력 및 기초 M/B 연산
# --------------------------------------------------------------------------
with main_tab1:
    with st.expander("📷 [AI Vision] 수기 실험 노트/기록지 사진으로 자동 입력 (클릭하여 열기)", expanded=False):
        col_img1, col_img2 = st.columns([2, 1])
        with col_img1:
            uploaded_note_img = st.file_uploader(
                "수기로 적은 실험 일지 사진 업로드 (JPG, PNG)", 
                type=["jpg", "jpeg", "png"],
                key="up_note_img"
            )
        with col_img2:
            st.write("")
            st.write("")
            if uploaded_note_img is not None:
                if st.button("🚀 사진 분석 및 수치 자동 입력", type="primary", use_container_width=True):
                    with st.spinner("Gemini AI가 수기 노트를 판독하고 있습니다..."):
                        img_bytes = uploaded_note_img.read()
                        parsed_data, err = parse_image_with_vision(img_bytes, doc_type="lab_note")
                        if err:
                            st.error(f"❌ 분석 실패: {err}")
                        elif parsed_data:
                            applied_cnt = 0
                            for k, v in parsed_data.items():
                                if v is not None and k in st.session_state:
                                    st.session_state[k] = float(v)
                                    applied_cnt += 1
                            st.success(f"🎉 판독 완료! 총 {applied_cnt}개 수치가 입력창에 자동 반영되었습니다.")
                            st.rerun()

    with st.expander("📝 이번 회차 실험 수치 입력 폼", expanded=True):
        col_in1, col_in2 = st.columns(2)
        with col_in1:
            st.markdown("#### [1. 투입 원료 및 반응 조건]")
            run_no = st.number_input("실험 회차 (Run No.)", min_value=1, value=int(st.session_state["run_no"]), step=1, key="run_no")
            li2co3_mass = st.number_input("Li₂CO₃ 투입량 (g)", value=float(st.session_state["li2co3_mass"]), format="%.2f", key="inp_li2co3_mass")
            li2co3_water = st.number_input("Li₂CO₃ 용매수 (g)", value=float(st.session_state["li2co3_water"]), format="%.1f", key="inp_li2co3_water")
            fresh_cao_mass = st.number_input("신품 CaO 투입량 (g)", value=float(st.session_state["fresh_cao_mass"]), format="%.2f", key="inp_fresh_cao_mass")
            recycled_cao_mass = st.number_input("재생 CaO 투입량 (g)", value=float(st.session_state["recycled_cao_mass"]), format="%.2f", key="inp_recycled_cao_mass")
            slurry_water = st.number_input("슬러리 조제수 (g)", value=float(st.session_state["slurry_water"]), format="%.1f", key="inp_slurry_water")
            temp_c = st.number_input("반응 온도 (℃)", value=float(st.session_state["temp_c"]), format="%.1f", key="inp_temp_c")
            time_h = st.number_input("반응 시간 (시간)", value=float(st.session_state["time_h"]), format="%.1f", key="inp_time_h")

        with col_in2:
            st.markdown("#### [2. 1차 여과 및 케이크 수세]")
            primary_filtrate_mass = st.number_input("1차 LiOH 여액 무게 (g)", value=float(st.session_state["primary_filtrate_mass"]), format="%.1f", key="inp_primary_filtrate_mass")
            primary_filtrate_sg = st.number_input("1차 여액 비중 (g/mL)", value=float(st.session_state["primary_filtrate_sg"]), format="%.3f", step=0.001, key="inp_primary_filtrate_sg")
            primary_filtrate_ph = st.number_input("1차 여액 pH", value=float(st.session_state["primary_filtrate_ph"]), format="%.2f", step=0.05, key="inp_primary_filtrate_ph")
            wet_cake_mass = st.number_input("1차 습케이크 무게 (g)", value=float(st.session_state["wet_cake_mass"]), format="%.1f", key="inp_wet_cake_mass")
            sample_wet = st.number_input("함수율 측정 샘플 습중량 (g)", value=float(st.session_state["sample_wet"]), format="%.1f", key="inp_sample_wet")
            sample_dry = st.number_input("함수율 측정 샘플 건중량 (g)", value=float(st.session_state["sample_dry"]), format="%.1f", key="inp_sample_dry")
            wash_sol_mass = st.number_input("회수된 수세액 무게 (g)", value=float(st.session_state["wash_sol_mass"]), format="%.1f", key="inp_wash_sol_mass")
            wash_sol_sg = st.number_input("수세액 비중 (g/mL)", value=float(st.session_state["wash_sol_sg"]), format="%.3f", step=0.001, key="inp_wash_sol_sg")
            wash_sol_ph = st.number_input("수세액 pH", value=float(st.session_state["wash_sol_ph"]), format="%.2f", step=0.05, key="inp_wash_sol_ph")

        st.divider()

        st.markdown("#### [3. CaCO₃ 소성(하소) 및 CaO 재생]")
        col_calc1, col_calc2 = st.columns(2)
        with col_calc1:
            test_dry_cake = st.number_input("소성 투입 건조케익 샘플 (g)", value=float(st.session_state["test_dry_cake"]), format="%.1f", key="inp_test_dry_cake")
            calcined_cao = st.number_input("소성 후 회수된 CaO (g)", value=float(st.session_state["calcined_cao"]), format="%.1f", key="inp_calcined_cao")
        with col_calc2:
            calc_temp = st.number_input("소성 온도 (℃)", value=float(st.session_state["calc_temp"]), format="%.1f", key="inp_calc_temp")
            calc_time = st.number_input("소성 시간 (시간)", value=float(st.session_state["calc_time"]), format="%.1f", key="inp_calc_time")

    n_li2co3 = li2co3_mass / MW_LI2CO3
    total_cao_in = fresh_cao_mass + recycled_cao_mass
    n_cao = (total_cao_in * 1.0) / MW_CAO
    limiting = "Li2CO3" if n_li2co3 <= n_cao else "CaO"
    excess_pct = (n_cao / n_li2co3 - 1.0) * 100
    theo_lioh_mass = (n_li2co3 * 2.0) * MW_LIOH

    total_in = li2co3_mass + li2co3_water + total_cao_in + slurry_water
    total_out = primary_filtrate_mass + wet_cake_mass
    loss_mass = total_in - total_out
    mass_closure = (total_out / total_in) * 100.0
    cake_moisture = (1.0 - (sample_dry / sample_wet)) * 100.0 if sample_wet > 0 else 0.0
    est_total_dry_solids = wet_cake_mass * (sample_dry / sample_wet) if sample_wet > 0 else 0.0

    loi_pct = ((test_dry_cake - calcined_cao) / test_dry_cake) * 100.0 if test_dry_cake > 0 else 0.0
    cao_yield_dry = (calcined_cao / test_dry_cake) * 100.0 if test_dry_cake > 0 else 0.0
    purity_caco3 = max(0.0, min(100.0, ((loi_pct/100.0) - 0.2432) / (0.4397 - 0.2432) * 100.0))
    pot_total_cao = est_total_dry_solids * (cao_yield_dry / 100.0)
    ca_loop_recovery = (pot_total_cao / total_cao_in) * 100.0 if total_cao_in > 0 else 0.0

    target_cao = 92.42
    fresh_makeup = max(0.0, target_cao - calcined_cao)

    st.subheader(f"📊 Run {run_no} 공정 기본 물질수지(M/B)")
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("M/B 정합성 (Closure)", f"{mass_closure:.2f} %", f"증발: {loss_mass:.1f}g")
    k2.metric("1차 케익 함수율", f"{cake_moisture:.2f} %", f"고형분 {est_total_dry_solids:.1f}g")
    k3.metric("소성 감율 (LOI)", f"{loi_pct:.2f} %", f"CaCO₃ 순도 ~{purity_caco3:.1f}%")
    k4.metric("Ca-Loop 원소 회수율", f"{ca_loop_recovery:.2f} %", f"재생잠재 {pot_total_cao:.1f}g")

# --------------------------------------------------------------------------
# TAB 2: 🧪 용액 ICP 분석
# --------------------------------------------------------------------------
with main_tab2:
    st.header("🧪 용액 ICP 분석 데이터 입력 및 회수율 계산")
    st.markdown("분석실 **성적서 사진(Gemini Vision)** 또는 **엑셀 파일**을 업로드하면 성분값이 자동 채워집니다.")

    with st.expander("📷 [AI Vision] ICP 성적서 / 기기 화면 사진으로 자동 입력 (클릭하여 열기)", expanded=False):
        col_icp_img1, col_icp_img2 = st.columns([2, 1])
        with col_icp_img1:
            uploaded_icp_img = st.file_uploader(
                "ICP 성적서 또는 기기 화면 사진 업로드 (JPG, PNG)", 
                type=["jpg", "jpeg", "png"],
                key="up_icp_img"
            )
        with col_icp_img2:
            st.write("")
            st.write("")
            if uploaded_icp_img is not None:
                if st.button("🚀 ICP 사진 판독 및 자동 입력", type="primary", use_container_width=True):
                    with st.spinner("Gemini AI가 성분 분석표를 판독하고 있습니다..."):
                        img_bytes = uploaded_icp_img.read()
                        parsed_icp, err = parse_image_with_vision(img_bytes, doc_type="icp_report")
                        if err:
                            st.error(f"❌ 분석 실패: {err}")
                        elif parsed_icp:
                            for k, v in parsed_icp.items():
                                if v is not None and k in st.session_state:
                                    st.session_state[k] = float(v)
                            st.success("🎉 ICP 성분값(Li, Ca, Na, Si, Mg, K) 판독 및 자동 반영 완료!")
                            st.rerun()

    with st.container():
        col_up1, col_up2 = st.columns([3, 1])
        with col_up1:
            uploaded_icp_file = st.file_uploader(
                "📂 또는 ICP 분석 엑셀/CSV 파일 업로드 (형태 A)", 
                type=["xlsx", "xls", "csv"]
            )
        with col_up2:
            st.write("")
            st.write("")
            df_template = pd.DataFrame({
                "시료명 (Sample)": ["1차 여액 (Primary Filtrate)", "수세액 (Wash Solution)"],
                "Li (mg/L)": [10500.0, 1400.0],
                "Ca (mg/L)": [120.0, 80.0],
                "Na (mg/L)": [45.0, 6.0],
                "Si (mg/L)": [8.5, 2.1],
                "Mg (mg/L)": [1.2, 0.3],
                "K (mg/L)": [15.0, 2.0]
            })
            tpl_buffer = io.BytesIO()
            with pd.ExcelWriter(tpl_buffer, engine='openpyxl') as writer:
                df_template.to_excel(writer, index=False, sheet_name="ICP_Analysis")
            tpl_buffer.seek(0)

            st.download_button(
                label="📥 표준 엑셀 양식 다운로드",
                data=tpl_buffer,
                file_name="ICP_Analysis_Template_FormA.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

        if uploaded_icp_file is not None:
            try:
                if uploaded_icp_file.name.endswith(".csv"):
                    df_up = pd.read_csv(uploaded_icp_file)
                else:
                    df_up = pd.read_excel(uploaded_icp_file)

                sample_col = None
                for col in df_up.columns:
                    if any(k in str(col).lower() for k in ["시료", "sample", "구분", "item", "name", "용액"]):
                        sample_col = col
                        break
                if sample_col is None:
                    sample_col = df_up.columns[0]

                row_1, row_w = None, None
                for idx, val in df_up[sample_col].astype(str).items():
                    v_clean = val.strip().lower()
                    if any(k in v_clean for k in ["1차", "여액", "primary", "1st", "filtrate"]): row_1 = idx
                    elif any(k in v_clean for k in ["수세", "세척", "wash"]): row_w = idx

                if row_1 is None and len(df_up) >= 1: row_1 = 0
                if row_w is None and len(df_up) >= 2: row_w = 1

                elem_mapping = {
                    "Li": ("icp_li_1", "icp_li_w"), "Ca": ("icp_ca_1", "icp_ca_w"),
                    "Na": ("icp_na_1", "icp_na_w"), "Si": ("icp_si_1", "icp_si_w"),
                    "Mg": ("icp_mg_1", "icp_mg_w"), "K":  ("icp_k_1", "icp_k_w")
                }

                matched_elems = []
                for col in df_up.columns:
                    c_clean = str(col).strip().upper()
                    for el, (k1, kw) in elem_mapping.items():
                        tokens = [t.strip("()[],._") for t in c_clean.split()]
                        first_tok = tokens[0] if tokens else ""
                        if first_tok == el.upper() or c_clean.startswith(el.upper()):
                            if row_1 is not None and not pd.isna(df_up.loc[row_1, col]):
                                st.session_state[k1] = float(df_up.loc[row_1, col])
                            if row_w is not None and not pd.isna(df_up.loc[row_w, col]):
                                st.session_state[kw] = float(df_up.loc[row_w, col])
                            matched_elems.append(el)
                            break

                matched_elems = list(set(matched_elems))
                st.success(f"🎉 엑셀 분석 완료! 매칭된 성분: **{', '.join(matched_elems)}**")
            except Exception as e:
                st.error(f"❌ 엑셀 파싱 오류: {e}")

    st.divider()

    st.markdown("### 1. ICP 분석 데이터 확인 및 수정 (단위: mg/L)")
    icp_col1, icp_col2 = st.columns(2)

    with icp_col1:
        st.markdown(f"#### 🔹 1차 여액 분석치 (부피: {primary_filtrate_mass/primary_filtrate_sg:.1f} mL)")
        icp_li_1 = st.number_input("Li 농도 (mg/L) - 1차 여액", value=float(st.session_state["icp_li_1"]), step=50.0, format="%.1f", key="icp_li_1")
        icp_ca_1 = st.number_input("Ca 농도 (mg/L) - 1차 여액", value=float(st.session_state["icp_ca_1"]), step=5.0, format="%.1f", key="icp_ca_1")
        icp_na_1 = st.number_input("Na 농도 (mg/L) - 1차 여액", value=float(st.session_state["icp_na_1"]), step=1.0, format="%.1f", key="icp_na_1")
        icp_si_1 = st.number_input("Si 농도 (mg/L) - 1차 여액", value=float(st.session_state["icp_si_1"]), step=0.5, format="%.1f", key="icp_si_1")
        icp_mg_1 = st.number_input("Mg 농도 (mg/L) - 1차 여액", value=float(st.session_state["icp_mg_1"]), step=0.1, format="%.1f", key="icp_mg_1")
        icp_k_1  = st.number_input("K 농도 (mg/L) - 1차 여액", value=float(st.session_state["icp_k_1"]), step=1.0, format="%.1f", key="icp_k_1")

    with icp_col2:
        st.markdown(f"#### 🔹 수세액 분석치 (부피: {wash_sol_mass/wash_sol_sg:.1f} mL)")
        icp_li_w = st.number_input("Li 농도 (mg/L) - 수세액", value=float(st.session_state["icp_li_w"]), step=50.0, format="%.1f", key="icp_li_w")
        icp_ca_w = st.number_input("Ca 농도 (mg/L) - 수세액", value=float(st.session_state["icp_ca_w"]), step=5.0, format="%.1f", key="icp_ca_w")
        icp_na_w = st.number_input("Na 농도 (mg/L) - 수세액", value=float(st.session_state["icp_na_w"]), step=1.0, format="%.1f", key="icp_na_w")
        icp_si_w = st.number_input("Si 농도 (mg/L) - 수세액", value=float(st.session_state["icp_si_w"]), step=0.5, format="%.1f", key="icp_si_w")
        icp_mg_w = st.number_input("Mg 농도 (mg/L) - 수세액", value=float(st.session_state["icp_mg_w"]), step=0.1, format="%.1f", key="icp_mg_w")
        icp_k_w  = st.number_input("K 농도 (mg/L) - 수세액", value=float(st.session_state["icp_k_w"]), step=1.0, format="%.1f", key="icp_k_w")

    st.divider()

    st.markdown("### 2. 성분별 질량 및 회수율 계산 결과")

    v1_L = (primary_filtrate_mass / primary_filtrate_sg) / 1000.0
    vw_L = (wash_sol_mass / wash_sol_sg) / 1000.0
    li_in_total_g = n_li2co3 * 2.0 * MW_LI

    elements = ["Li", "Ca", "Na", "Si", "Mg", "K"]
    conc_1 = [icp_li_1, icp_ca_1, icp_na_1, icp_si_1, icp_mg_1, icp_k_1]
    conc_w = [icp_li_w, icp_ca_w, icp_na_w, icp_si_w, icp_mg_w, icp_k_w]

    mass_1_g = [c * v1_L / 1000.0 for c in conc_1]
    mass_w_g = [c * vw_L / 1000.0 for c in conc_w]
    mass_total_g = [m1 + mw for m1, mw in zip(mass_1_g, mass_w_g)]

    li_rec_1_pct = (mass_1_g[0] / li_in_total_g) * 100.0 if li_in_total_g > 0 else 0.0
    li_rec_w_pct = (mass_w_g[0] / li_in_total_g) * 100.0 if li_in_total_g > 0 else 0.0
    total_li_rec_pct = li_rec_1_pct + li_rec_w_pct
    li_loss_pct = max(0.0, 100.0 - total_li_rec_pct)
    lioh_equiv_g_l = icp_li_1 * (MW_LIOH / MW_LI) / 1000.0

    rc1, rc2, rc3, rc4 = st.columns(4)
    rc1.metric("🎯 총 Li 회수율", f"{total_li_rec_pct:.2f} %", f"총 회수: {mass_total_g[0]:.2f}g / 투입: {li_in_total_g:.2f}g")
    rc2.metric("1차 여액 회수율", f"{li_rec_1_pct:.2f} %", f"회수 질량: {mass_1_g[0]:.2f}g")
    rc3.metric("수세액 회수율", f"{li_rec_w_pct:.2f} %", f"회수 질량: {mass_w_g[0]:.2f}g")
    rc4.metric("케이크 잔류/손실률", f"{li_loss_pct:.2f} %", f"미회수: {li_in_total_g - mass_total_g[0]:.2f}g", delta_color="inverse")

    df_icp_summary = pd.DataFrame({
        "원소 (Element)": elements,
        "1차 여액 농도 (mg/L)": conc_1,
        "1차 여액 회수량 (g)": [round(x, 4) for x in mass_1_g],
        "수세액 농도 (mg/L)": conc_w,
        "수세액 회수량 (g)": [round(x, 4) for x in mass_w_g],
        "총 용출 질량 (g)": [round(x, 4) for x in mass_total_g],
        "Li 대비 불순물 비율 (wt ppm)": [
            round((mt / mass_total_g[0]) * 1e6, 1) if el != "Li" and mass_total_g[0] > 0 else "-" 
            for el, mt in zip(elements, mass_total_g)
        ]
    })

    st.markdown("##### 📋 다성분(Li, Ca, Na, Si, Mg, K) 분석 및 불순물 비율 요약표")
    st.dataframe(
        df_icp_summary.style.format({
            "1차 여액 농도 (mg/L)": "{:,.1f}", "1차 여액 회수량 (g)": "{:.4f}",
            "수세액 농도 (mg/L)": "{:,.1f}", "수세액 회수량 (g)": "{:.4f}",
            "총 용출 질량 (g)": "{:.4f}"
        }),
        use_container_width=True
    )

    st.markdown("---")
    col_sv1, col_sv2 = st.columns([2, 1])
    with col_sv1:
        save_clicked = st.button("💾 이 분석 결과를 트렌드 DB에 저장 (및 엑셀 리포트 자동 메일 발송)", type="primary", use_container_width=True)
    with col_sv2:
        st.session_state.auto_email_on_save = st.checkbox("저장 시 메일 자동 발송 켜기", value=st.session_state.auto_email_on_save)

    if save_clicked:
        new_row = {
            "회차 (Run)": int(run_no),
            "구분": "실측치 (Actual)",
            "Li 회수율 (%)": round(total_li_rec_pct, 2), 
            "1차여액 Li농도 (mg/L)": round(icp_li_1, 1),
            "1차여액 LiOH농도 (g/L)": round(lioh_equiv_g_l, 2),
            "M/B 닫힘율 (%)": round(mass_closure, 2), 
            "하소 감율 LOI (%)": round(loi_pct, 2), 
            "CaO 활성도 (%)": 100.0,
            "신품 CaO 보충량 (g)": round(fresh_makeup, 2)
        }
        st.session_state.history = st.session_state.history[st.session_state.history["회차 (Run)"] != run_no]
        st.session_state.history = pd.concat([st.session_state.history, pd.DataFrame([new_row])]).sort_values("회차 (Run)").reset_index(drop=True)
        st.success(f"✅ Run {run_no} ICP 분석 데이터가 트렌드 DB에 영구 등록되었습니다!")

        if st.session_state.auto_email_on_save:
            ok, msg_res = send_email_report(
                run_num=run_no, mass_cls=mass_closure, loss_m=loss_mass,
                li_rec_tot=total_li_rec_pct, li_rec_1=li_rec_1_pct, li_rec_w=li_rec_w_pct,
                lioh_conc=lioh_equiv_g_l, li_1=icp_li_1, ca_1=icp_ca_1, na_1=icp_na_1,
                si_1=icp_si_1, mg_1=icp_mg_1, k_1=icp_k_1, loi=loi_pct,
                purity=purity_caco3, makeup=fresh_makeup, cao_rec=calcined_cao,
                df_icp_tbl=df_icp_summary, df_sim_tbl=None, is_auto=True
            )
            if ok:
                st.toast(f"📧 리포트 메일 자동 발송 완료!", icon="🎉")
                st.success(f"📧 **[자동 발송 성공]** {msg_res}")
            else:
                st.warning(f"⚠️ **[자동 발송 미완료]** {msg_res}")

# --------------------------------------------------------------------------
# TAB 3: 📈 회차별 트렌드 & 거동예측
# --------------------------------------------------------------------------
with main_tab3:
    st.header("📈 $n$회차 트렌드 시각화 & 거동예측 시뮬레이터")
    
    with st.expander("⚙️ 거동예측 시뮬레이션 파라미터 설정 (What-If Simulation)", expanded=True):
        col_p1, col_p2, col_p3 = st.columns(3)
        with col_p1:
            target_max_run = st.slider("예측 시뮬레이션 목표 회차", min_value=3, max_value=20, value=10, step=1)
            sintering_decay = st.slider("회차당 CaO 소결 활성도 감쇄율 (%)", min_value=0.5, max_value=8.0, value=3.5, step=0.1)
        with col_p2:
            sim_temp = st.number_input("가상 반응 온도 (℃)", min_value=60.0, max_value=95.0, value=80.0, step=1.0)
            sim_time = st.number_input("가상 반응 시간 (h)", min_value=1.0, max_value=5.0, value=2.0, step=0.5)
        with col_p3:
            wash_ratio = st.slider("수세수 투입 배수 (케이크 대비)", min_value=1.0, max_value=5.0, value=3.0, step=0.5)
            fresh_makeup_mode = st.selectbox("CaO 보충 방식", ["고정량 보충 (전회차 감량분 100% Make-up)", "신품 100% 교체 (Purge)"])

    sim_rows = []
    base_li_conc = icp_li_1 if 'icp_li_1' in locals() else 10500.0
    base_recovery = total_li_rec_pct if 'total_li_rec_pct' in locals() and total_li_rec_pct > 0 else 95.80

    temp_factor = 1.0 + (sim_temp - 80.0) * 0.004
    time_factor = 1.0 + (sim_time - 2.0) * 0.03

    for r in range(1, target_max_run + 1):
        matched_actual = st.session_state.history[st.session_state.history["회차 (Run)"] == r]
        
        if not matched_actual.empty:
            row = matched_actual.iloc[0].to_dict()
            sim_rows.append(row)
        else:
            activity = 100.0 * ((1.0 - (sintering_decay / 100.0)) ** (r - 1))
            eff_activity = min(100.0, activity * temp_factor * time_factor)
            
            pred_rec = max(70.0, min(99.0, base_recovery * (eff_activity / 100.0)))
            pred_li_conc = max(7000.0, base_li_conc * (pred_rec / base_recovery))
            pred_lioh_conc = round(pred_li_conc * (MW_LIOH / MW_LI) / 1000.0, 2)
            pred_loi = max(35.0, 41.13 - (r - 1) * 0.4)
            pred_makeup = 68.52 if fresh_makeup_mode.startswith("고정량") else 92.42

            sim_rows.append({
                "회차 (Run)": r,
                "구분": "AI 예측치 (Simulated)",
                "Li 회수율 (%)": round(pred_rec, 2),
                "1차여액 Li농도 (mg/L)": round(pred_li_conc, 1),
                "1차여액 LiOH농도 (g/L)": round(pred_lioh_conc, 2),
                "M/B 닫힘율 (%)": round(max(90.0, 95.06 - (r - 1) * 0.2), 2),
                "하소 감율 LOI (%)": round(pred_loi, 2),
                "CaO 활성도 (%)": round(eff_activity, 1),
                "신품 CaO 보충량 (g)": round(pred_makeup, 2)
            })

    df_simulation = pd.DataFrame(sim_rows)

    st.markdown("---")
    st.subheader("📊 회차별 인터랙티브 X-Y 트렌드 그래프")

    col_ctrl1, col_ctrl2 = st.columns([1, 2])
    with col_ctrl1:
        y_axis_metric = st.selectbox(
            "📌 Y축에 표시할 지표를 선택하세요:",
            [
                "Li 회수율 (%)", "1차여액 LiOH농도 (g/L)", "1차여액 Li농도 (mg/L)", 
                "CaO 활성도 (%)", "하소 감율 LOI (%)", "신품 CaO 보충량 (g)"
            ]
        )

    chart_df = df_simulation.set_index("회차 (Run)")[[y_axis_metric]]
    st.line_chart(chart_df, height=380, use_container_width=True)

    st.markdown("##### 📋 회차별 실측치 & 예측 시뮬레이션 상세 데이터 테이블")
    st.dataframe(
        df_simulation.style.format({
            "Li 회수율 (%)": "{:.2f} %", "1차여액 Li농도 (mg/L)": "{:,.1f} mg/L",
            "1차여액 LiOH농도 (g/L)": "{:.2f} g/L", "M/B 닫힘율 (%)": "{:.2f} %",
            "하소 감율 LOI (%)": "{:.2f} %", "CaO 활성도 (%)": "{:.1f} %",
            "신품 CaO 보충량 (g)": "{:.2f} g"
        }),
        use_container_width=True
    )

# --------------------------------------------------------------------------
# TAB 4: 💬 AI 공정 대화창
# --------------------------------------------------------------------------
with main_tab4:
    st.header("💬 AI 공정 엔지니어와 대화하기")
    st.caption("현재 실험 수치, ICP 분석 결과, $n$회차 시뮬레이션을 바탕으로 실시간 질의응답을 진행합니다.")

    for msg in st.session_state.chat_messages:
        with st.chat_message(msg["role"]):
            st.markdown(msg["content"])

    if user_prompt := st.chat_input("질문을 입력하세요 (예: Ca 농도가 120ppm이면 제품 순도에 영향이 커?)"):
        st.session_state.chat_messages.append({"role": "user", "content": user_prompt})
        with st.chat_message("user"):
            st.markdown(user_prompt)

        prompt_low = user_prompt.lower()
        if "ca" in prompt_low or "칼슘" in prompt_low or "불순물" in prompt_low:
            ai_reply = f"현재 1차 여액 내 Ca 농도는 **{icp_ca_1} mg/L**입니다. 배터리급 수산화리튬(LH) 스펙 대비 높으므로, 증발 농축/결정화 전에 CO₂ 가스를 불어넣어 침전 분리하는 **2차 탄산화(Carbonation) 탈칼슘 공정**을 거치는 것을 권장합니다."
        elif "ph" in prompt_low or "수세" in prompt_low or "세척" in prompt_low:
            ai_reply = f"수세액 pH가 **{wash_sol_ph}**로 높은 이유는 1차 감압 여과 후 케이크 기공 내 LiOH 고농도 액이 갇혀 있었기 때문입니다. 3배수 수세를 통해 총 {mass_w_g[0]:.2f}g의 Li(회수율 기여도 {li_rec_w_pct:.2f}%)을 회수하였습니다."
        elif "소결" in prompt_low or "sintering" in prompt_low or "퍼지" in prompt_low:
            ai_reply = f"1000℃ 고온 하소가 반복되면 활성도가 회차당 약 {sintering_decay}%씩 저하됩니다. 회수율을 90% 이상 유지하려면 약 6~7회차 시점에 재생 CaO 일부를 퍼지(Purge)하고 신품 CaO로 리프레시해 주시는 것이 좋습니다."
        else:
            ai_reply = f"Run {run_no}의 총 Li 회수율은 **{total_li_rec_pct:.2f}%** (1차 여액 {li_rec_1_pct:.1f}% + 수세액 {li_rec_w_pct:.1f}%)로 산출되었습니다. 추가로 확인하고 싶은 반응 변수나 거동 조건이 있으신가요?"

        st.session_state.chat_messages.append({"role": "assistant", "content": ai_reply})
        with st.chat_message("assistant"):
            st.markdown(ai_reply)

# --------------------------------------------------------------------------
# TAB 5: 📧 리포트 메일 발송 현황 & 연결 테스트
# --------------------------------------------------------------------------
with main_tab5:
    st.header("📧 리포트 메일 발송 현황 & 계정 설정")
    st.caption("실험 데이터 저장 시 자동으로 발송된 이메일 현황 목록을 모니터링하고 발송 설정을 관리합니다.")

    st.markdown("### 📋 실시간 메일 발송 이력 현황")
    if st.session_state.email_logs:
        df_logs = pd.DataFrame(st.session_state.email_logs)
        success_cnt = sum(1 for log in st.session_state.email_logs if "성공" in log["발송 상태"])
        fail_cnt = len(st.session_state.email_logs) - success_cnt

        m_col1, m_col2, m_col3 = st.columns(3)
        m_col1.metric("총 발송 시도", f"{len(st.session_state.email_logs)} 건")
        m_col2.metric("발송 성공", f"{success_cnt} 건")
        m_col3.metric("발송 실패 / 보류", f"{fail_cnt} 건", delta_color="inverse")

        st.dataframe(df_logs, use_container_width=True)

        if st.button("🗑️ 발송 이력 초기화", use_container_width=False):
            st.session_state.email_logs = []
            st.rerun()
    else:
        st.info("ℹ️ 아직 발송된 메일 이력이 없습니다. 2번 탭에서 [💾 트렌드 DB에 저장]을 누르면 엑셀 리포트가 자동 발송되고 여기에 기록됩니다.")

    st.divider()

    st.markdown("### ⚙️ 발송 계정 및 수신자 설정")
    col_cfg1, col_cfg2 = st.columns(2)

    with col_cfg1:
        st.session_state.email_recipients = st.text_input(
            "📬 수신자 이메일 목록 (쉼표로 구분)", 
            value=st.session_state.email_recipients,
            help="예: user1@company.com, user2@company.com"
        )
        st.session_state.email_sender = st.text_input(
            "📤 발신자 이메일 주소 (예: myname@gmail.com)", 
            value=st.session_state.email_sender
        )
        st.session_state.email_password = st.text_input(
            "🔑 발신자 앱 비밀번호 (16자리)", 
            value=st.session_state.email_password,
            type="password",
            help="Gmail: Google 계정 보안 > 2단계 인증 > 앱 비밀번호"
        )

    with col_cfg2:
        st.session_state.smtp_server = st.text_input(
            "🌐 SMTP 서버 호스트", 
            value=st.session_state.smtp_server
        )
        st.session_state.smtp_port = st.number_input(
            "🔌 SMTP 포트 번호", 
            value=int(st.session_state.smtp_port),
            step=1
        )
        st.write("")
        st.session_state.auto_email_on_save = st.toggle(
            "⚡ 실험값 저장 시 엑셀 리포트 자동 발송 활성화", 
            value=st.session_state.auto_email_on_save
        )

    st.markdown("---")
    st.markdown("#### 🔍 계정 연결 1초 진단 테스트")
    col_t1, col_t2 = st.columns(2)
    with col_t1:
        if st.button("🔌 SMTP 계정 연결 즉시 테스트", use_container_width=True):
            test_sender = st.session_state.email_sender.strip()
            test_pw = st.session_state.email_password.strip().replace(" ", "")
            test_host = st.session_state.smtp_server.strip()
            test_port = int(st.session_state.smtp_port)

            if not test_sender or not test_pw:
                st.error("❌ 발신자 이메일과 비밀번호를 입력한 후 테스트를 눌러주세요.")
            else:
                with st.spinner("메일 서버 연결 및 로그인 인증 테스트 중..."):
                    try:
                        if test_port == 465:
                            s = smtplib.SMTP_SSL(test_host, test_port, timeout=10)
                        else:
                            s = smtplib.SMTP(test_host, test_port, timeout=10)
                            s.starttls()
                        s.login(test_sender, test_pw)
                        s.quit()
                        st.success(f"🎉 **[인증 성공!]** `{test_sender}` 계정으로 메일 서버에 정상 로그인되었습니다!")
                    except Exception as err:
                        st.error(f"❌ **[인증 실패]** 상세 원인: {err}")

    with col_t2:
        if st.button("📨 현재 회차 리포트 수동 발송", type="primary", use_container_width=True):
            ok, msg_res = send_email_report(
                run_num=run_no, mass_cls=mass_closure, loss_m=loss_mass,
                li_rec_tot=total_li_rec_pct, li_rec_1=li_rec_1_pct, li_rec_w=li_rec_w_pct,
                lioh_conc=lioh_equiv_g_l, li_1=icp_li_1, ca_1=icp_ca_1, na_1=icp_na_1,
                si_1=icp_si_1, mg_1=icp_mg_1, k_1=icp_k_1, loi=loi_pct,
                purity=purity_caco3, makeup=fresh_makeup, cao_rec=calcined_cao,
                df_icp_tbl=df_icp_summary, df_sim_tbl=df_simulation, is_auto=False
            )
            if ok:
                st.success(f"🎉 {msg_res}")
                st.rerun()
            else:
                st.error(f"❌ {msg_res}")
                st.rerun()
